"""Lectura y escritura Excel (openpyxl). Un archivo por proveedor."""
from __future__ import annotations

import re
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from database import normalize_reference

# Encabezados típicos (español / inglés), minúsculas para matching
REF_KEYS = (
    "referencia",
    "ref",
    "codigo",
    "código",
    "sku",
    "articulo",
    "artículo",
    "producto",
    "item",
)
COST_KEYS = (
    "costo",
    "precio",
    "pvp",
    "p.v.p",
    "importe",
    "price",
    "cost",
    "tarifa",
    "valor",
)
DESC_KEYS = (
    "descripcion",
    "descripción",
    "nombre",
    "desc",
    "description",
)


def _norm_header(cell: Any) -> str:
    if cell is None:
        return ""
    t = str(cell).strip().lower()
    t = re.sub(r"\s+", " ", t)
    return t


def _header_score(header: str, keys: tuple[str, ...]) -> float:
    if not header:
        return 0.0
    best = 0.0
    for k in keys:
        if header == k:
            best = max(best, 1.0)
        elif k in header or header in k:
            best = max(best, 0.6)
    return best


def detect_columns(headers: list[str]) -> tuple[int | None, int | None, int | None]:
    """Devuelve índices 0-based (ref, cost, desc opcional)."""
    scored_ref: list[tuple[float, int]] = []
    scored_cost: list[tuple[float, int]] = []
    scored_desc: list[tuple[float, int]] = []
    for i, h in enumerate(headers):
        nh = _norm_header(h)
        scored_ref.append((_header_score(nh, REF_KEYS), i))
        scored_cost.append((_header_score(nh, COST_KEYS), i))
        scored_desc.append((_header_score(nh, DESC_KEYS), i))
    scored_ref.sort(reverse=True)
    scored_cost.sort(reverse=True)
    scored_desc.sort(reverse=True)
    ref_i = scored_ref[0][1] if scored_ref and scored_ref[0][0] >= 0.6 else None
    cost_i = scored_cost[0][1] if scored_cost and scored_cost[0][0] >= 0.6 else None
    desc_i = (
        scored_desc[0][1]
        if scored_desc and scored_desc[0][0] >= 0.6 and scored_desc[0][1] not in (ref_i, cost_i)
        else None
    )
    # Si hay dos columnas "precio", evitar usar la misma para ref y cost
    if ref_i is not None and cost_i is not None and ref_i == cost_i:
        for _, idx in scored_cost:
            if idx != ref_i:
                cost_i = idx
                break
    return ref_i, cost_i, desc_i


def parse_supplier_excel(path: Path) -> tuple[list[tuple[str, str, str | None, float, str | None]], str]:
    """
    Retorna (filas, nombre_proveedor_sugerido).
    filas: (reference_raw, reference_norm, description, cost, extra)
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], path.stem
    headers = [str(h) if h is not None else "" for h in header_row]
    ref_i, cost_i, desc_i = detect_columns(headers)
    wb.close()
    if ref_i is None or cost_i is None:
        raise ValueError(
            "No se detectaron columnas de referencia y costo. "
            "Use encabezados como Referencia/Código y Costo/Precio."
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    by_ref: dict[str, tuple[str, str, str | None, float, str | None]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ref_cell = row[ref_i] if ref_i < len(row) else None
        cost_cell = row[cost_i] if cost_i < len(row) else None
        desc_cell = row[desc_i] if desc_i is not None and desc_i < len(row) else None
        ref_raw = str(ref_cell).strip() if ref_cell is not None and str(ref_cell).strip() else ""
        if not ref_raw:
            continue
        ref_norm = normalize_reference(ref_raw)
        desc = None
        if desc_cell is not None and str(desc_cell).strip():
            desc = str(desc_cell).strip()[:500]
        try:
            if cost_cell is None:
                continue
            cost = float(cost_cell)
        except (TypeError, ValueError):
            continue
        by_ref[ref_norm] = (ref_raw, ref_norm, desc, cost, None)
    wb.close()
    out = list(by_ref.values())
    return out, path.stem


def export_comparison_excel(
    path: Path,
    reference: str,
    sale_price: float | None,
    lines: list[dict[str, Any]],
    explanation: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativa"
    ws.append(["Referencia buscada", reference])
    if sale_price is not None:
        ws.append(["Precio de venta (manual)", sale_price])
    ws.append([])
    headers = [
        "Orden",
        "Proveedor",
        "Referencia",
        "Descripción",
        "Costo proveedor",
    ]
    if sale_price is not None:
        headers.extend(["Ganancia por unidad", "Margen %"])
    ws.append(headers)
    for idx, line in enumerate(lines, start=1):
        row = [
            idx,
            line.get("supplier_name", ""),
            line.get("reference_raw", ""),
            line.get("description") or "",
            line.get("cost"),
        ]
        if sale_price is not None:
            c = float(line.get("cost") or 0)
            gain = sale_price - c
            margin = (gain / sale_price * 100.0) if sale_price else 0.0
            row.extend([round(gain, 4), round(margin, 2)])
        ws.append(row)
    ws.append([])
    ws.append(["Explicación recomendación"])
    for para in explanation.split("\n"):
        ws.append([para])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    wb.save(path)


def export_full_catalog(path: Path, conn: sqlite3.Connection) -> None:
    """Exporta todas las filas de la base (solo usuarios autenticados deben llamar esto)."""
    cur = conn.execute(
        """
        SELECT s.name AS supplier_name, p.reference_raw, p.description, p.cost,
               p.source_file, p.imported_at
        FROM price_rows p
        JOIN suppliers s ON s.id = p.supplier_id
        ORDER BY s.name COLLATE NOCASE, p.reference_norm COLLATE NOCASE
        """
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo"
    ws.append(
        [
            "Proveedor",
            "Referencia",
            "Descripción",
            "Costo",
            "Archivo origen",
            "Importado (UTC)",
        ]
    )
    for r in cur.fetchall():
        ws.append(
            [
                r["supplier_name"],
                r["reference_raw"],
                r["description"] or "",
                r["cost"],
                r["source_file"] or "",
                r["imported_at"] or "",
            ]
        )
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    wb.save(path)
